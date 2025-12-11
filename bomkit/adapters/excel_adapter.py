import openpyxl
from pathlib import Path

class ExcelAdapter:
    def can_handle(self, file_path):
        return Path(file_path).suffix.lower() in [".xlsx", ".xls"]

    def read(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        rows = []
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            rows.append(row_dict)

        return rows