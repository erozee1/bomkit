from .normalizer import BomNormalizer
from .schema import STANDARD_HEADERS
from typing import List, Dict, Any, Optional
from pathlib import Path
import csv
import json
import openpyxl


class BomParser:
    """Parser for Bill of Materials files with normalization support."""
    
    def __init__(self, normalize: bool = True):
        """Initialize the BOM parser.
        
        Args:
            normalize: If True, normalize rows to standard template (default: True)
        """
        self.adapters = []
        self.normalizer = BomNormalizer() if normalize else None
    
    def register_adapter(self, adapter):
        """Register a file adapter for parsing.
        
        Args:
            adapter: Adapter instance with can_handle() and read() methods
        """
        self.adapters.append(adapter)

    def parse(self, file_path: str, normalize: Optional[bool] = None) -> List[Dict[str, Any]]:
        """Parse a BOM file and optionally normalize to standard template.
        
        Args:
            file_path: Path to the BOM file
            normalize: Override default normalization setting (optional)
            
        Returns:
            List of dictionaries representing BOM rows.
            If normalization is enabled, rows use standard column names.
            Otherwise, returns raw rows with original column names.
            
        Raises:
            ValueError: If no adapter is found for the file
        """
        # Find appropriate adapter
        adapter = None
        for a in self.adapters:
            if a.can_handle(file_path):
                adapter = a
                break
        
        if adapter is None:
            raise ValueError(f"No adapter found for {file_path}")
        
        # Read raw rows
        raw_rows = adapter.read(file_path)
        
        # Normalize if requested
        should_normalize = normalize if normalize is not None else (self.normalizer is not None)
        if should_normalize and self.normalizer:
            return self.normalizer.normalize(raw_rows)
        
        return raw_rows
    
    def get_standard_template(self) -> List[str]:
        """Get the standard BOM template headers.
        
        Returns:
            List of standard header names
        """
        if self.normalizer:
            return self.normalizer.get_standard_template()
        return []
    
    def get_mapping_report(self, file_path: str) -> Dict[str, Any]:
        """Get a report of how columns from a file map to standard headers.
        
        Args:
            file_path: Path to the BOM file
            
        Returns:
            Dictionary with mapping information including mapped and unmapped columns
        """
        if not self.normalizer:
            return {"error": "Normalizer not initialized"}
        
        # Parse without normalization to get raw rows
        adapter = None
        for a in self.adapters:
            if a.can_handle(file_path):
                adapter = a
                break
        
        if adapter is None:
            raise ValueError(f"No adapter found for {file_path}")
        
        raw_rows = adapter.read(file_path)
        return self.normalizer.get_mapping_report(raw_rows)
    
    def export(self, data: List[Dict[str, Any]], output_path: str, format: Optional[str] = None) -> str:
        """Export normalized BOM data to a file.
        
        Args:
            data: List of dictionaries representing normalized BOM rows
            output_path: Path where the file should be saved
            format: Output format ('csv', 'excel', 'json', or None for auto-detect from extension)
            
        Returns:
            Path to the exported file
            
        Raises:
            ValueError: If format is not supported or data is empty
        """
        if not data:
            raise ValueError("Cannot export empty data")
        
        output_path = Path(output_path)
        
        # Auto-detect format from extension if not provided
        if format is None:
            suffix = output_path.suffix.lower()
            if suffix in ['.csv', '.tsv']:
                format = 'csv'
            elif suffix in ['.xlsx', '.xls']:
                format = 'excel'
            elif suffix == '.json':
                format = 'json'
            else:
                # Default to CSV if extension is not recognized
                format = 'csv'
                output_path = output_path.with_suffix('.csv')
        
        format = format.lower()
        
        # Get headers - use standard template if data appears normalized, otherwise use keys from first row
        if data and all(key in STANDARD_HEADERS for key in data[0].keys() if data[0]):
            headers = STANDARD_HEADERS
        else:
            # Use all unique keys from all rows
            all_keys = set()
            for row in data:
                all_keys.update(row.keys())
            headers = sorted(all_keys)
        
        # Export based on format
        if format == 'csv':
            self._export_csv(data, output_path, headers)
        elif format == 'excel':
            self._export_excel(data, output_path, headers)
        elif format == 'json':
            self._export_json(data, output_path)
        else:
            raise ValueError(f"Unsupported export format: {format}. Supported formats: csv, excel, json")
        
        return str(output_path)
    
    def _export_csv(self, data: List[Dict[str, Any]], output_path: Path, headers: List[str]) -> None:
        """Export data to CSV file."""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            for row in data:
                # Ensure all headers are present in row
                complete_row = {header: row.get(header, '') for header in headers}
                writer.writerow(complete_row)
    
    def _export_excel(self, data: List[Dict[str, Any]], output_path: Path, headers: List[str]) -> None:
        """Export data to Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(output_path)
    
    def _export_json(self, data: List[Dict[str, Any]], output_path: Path) -> None:
        """Export data to JSON file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    def parse_and_export(self, input_path: str, output_path: str, 
                        normalize: Optional[bool] = None, 
                        format: Optional[str] = None) -> str:
        """Parse a BOM file and export it to another format.
        
        Convenience method that combines parse() and export().
        
        Args:
            input_path: Path to the input BOM file
            output_path: Path where the exported file should be saved
            normalize: Whether to normalize data (defaults to parser's normalize setting)
            format: Output format ('csv', 'excel', 'json', or None for auto-detect)
            
        Returns:
            Path to the exported file
        """
        # Parse the input file
        data = self.parse(input_path, normalize=normalize)
        
        # Export to output file
        return self.export(data, output_path, format=format)